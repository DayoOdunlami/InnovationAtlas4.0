#!/usr/bin/env python3
"""Stage 2.6 — Testbed inventory ingestion.

Source: StaggingFiles/testbeds_metadata_augmented_v6 - Copy.xlsx
Target: cicerone_kb.testbeds

Column mapping (semantic match where names differ):
  spreadsheet column          → cicerone_kb.testbeds column
  -----------------------------------------------------
  Sector(s)                   → sector
  Location                    → location
  Access model                → access_model
  Operator(s)                 → operator
  Purpose (what you can test) → what_can_be_tested
  DSIT cluster                → dsit_cluster
  Confidence_score            → confidence_score
  (full row, all 15 columns)  → raw (jsonb)
  row position (1-indexed)    → row_number

Embedding text: "{sector} — {location} — {what_can_be_tested}".
text-embedding-3-small (1536).  pragma: allowlist secret

After ingestion the xlsx is moved to StaggingFiles/_processed/.

Run:
  python3 .tmp/ingest_cicerone_testbeds.py
"""

from __future__ import annotations

import json
import os
import shutil
import sys
from pathlib import Path
from typing import Any

import openpyxl
import psycopg2
import psycopg2.extras

_REPO_ROOT = Path(__file__).resolve().parents[1]
_DOTENV = _REPO_ROOT / ".env"
if _DOTENV.exists():
    for line in _DOTENV.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        v = v.strip().strip('"').strip("'")
        os.environ.setdefault(k.strip(), v)

import openai

EMBED_MODEL = "text-embedding-3-small"  # pragma: allowlist secret
SRC_PATH = _REPO_ROOT / "StaggingFiles" / "testbeds_metadata_augmented_v6 - Copy.xlsx"
PROCESSED_DIR = _REPO_ROOT / "StaggingFiles" / "_processed"

# Map spreadsheet headers (case-insensitive) → testbed column.
HEADER_MAP = {
    "sector(s)": "sector",
    "location": "location",
    "access model": "access_model",
    "operator(s)": "operator",
    "purpose (what you can test)": "what_can_be_tested",
    "dsit cluster": "dsit_cluster",
    "confidence_score": "confidence_score",
}


def vec_literal(emb: list[float]) -> str:
    return "[" + ",".join(f"{x:.7f}" for x in emb) + "]"


def main() -> int:
    pg_url = os.environ["POSTGRES_URL"].split("?")[0]
    db = psycopg2.connect(pg_url, sslmode="require")
    db.autocommit = False
    cur = db.cursor()
    embed_client = openai.OpenAI(api_key=os.environ["OPENAI_API_KEY"])

    if not SRC_PATH.exists():
        print(f"FATAL: source xlsx missing: {SRC_PATH}", file=sys.stderr)
        return 2

    print(f"Loading {SRC_PATH.name}...", flush=True)
    wb = openpyxl.load_workbook(SRC_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("FATAL: empty sheet", file=sys.stderr)
        return 3

    headers_raw = [(c if c is not None else "") for c in rows[0]]
    headers = [str(h).strip() for h in headers_raw]
    headers_lc = [h.lower() for h in headers]

    col_index: dict[str, int] = {}
    for tb_col in HEADER_MAP.values():
        col_index[tb_col] = -1
    mapping_log: list[str] = []
    for i, hl in enumerate(headers_lc):
        if hl in HEADER_MAP:
            tb = HEADER_MAP[hl]
            col_index[tb] = i
            mapping_log.append(f"  '{headers[i]}' (col {i}) -> {tb}")

    print("Column mapping:", flush=True)
    for line in mapping_log:
        print(line, flush=True)
    missing = [k for k, v in col_index.items() if v < 0]
    if missing:
        print(f"WARNING: unmapped target columns: {missing}", flush=True)

    # Idempotency: clear existing rows.
    cur.execute("DELETE FROM cicerone_kb.testbeds")

    inserted = 0
    skipped_blank = 0
    sample_rows: list[dict[str, Any]] = []

    for r_idx, row in enumerate(rows[1:], start=1):
        # Build raw jsonb dict from headers -> values.
        raw_dict = {}
        for i, h in enumerate(headers):
            v = row[i] if i < len(row) else None
            if isinstance(v, (int, float, str)) or v is None:
                raw_dict[h] = v
            else:
                raw_dict[h] = str(v)

        def get(col: str) -> Any:
            i = col_index[col]
            if i < 0:
                return None
            v = row[i] if i < len(row) else None
            if isinstance(v, str):
                v = v.strip() or None
            return v

        sector = get("sector")
        location = get("location")
        access_model = get("access_model")
        operator = get("operator")
        what_can_be_tested = get("what_can_be_tested")
        dsit_cluster = get("dsit_cluster")
        conf = get("confidence_score")
        try:
            confidence_score = float(conf) if conf is not None else None
        except (TypeError, ValueError):
            confidence_score = None

        # Skip blank rows.
        if not any([sector, location, access_model, operator, what_can_be_tested]):
            skipped_blank += 1
            continue

        embed_text_parts = [
            str(sector or "").strip(),
            str(location or "").strip(),
            str(what_can_be_tested or "").strip(),
        ]
        embed_text = " — ".join(p for p in embed_text_parts if p)
        if not embed_text:
            embed_text = json.dumps(raw_dict, default=str)[:2000]

        emb = (
            embed_client.embeddings.create(input=embed_text, model=EMBED_MODEL)
            .data[0]
            .embedding
        )

        cur.execute(
            """
            INSERT INTO cicerone_kb.testbeds (
                row_number, sector, location, access_model, operator,
                what_can_be_tested, dsit_cluster, confidence_score,
                raw, description_embedding
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, %s::vector)
            """,
            (
                r_idx,
                sector,
                location,
                access_model,
                operator,
                what_can_be_tested,
                dsit_cluster,
                confidence_score,
                json.dumps(raw_dict, default=str),
                vec_literal(emb),
            ),
        )
        inserted += 1
        if len(sample_rows) < 5:
            sample_rows.append(
                {
                    "row_number": r_idx,
                    "sector": sector,
                    "location": location,
                    "what_can_be_tested": (what_can_be_tested or "")[:120],
                }
            )

    db.commit()
    print(f"\nINSERTED rows: {inserted}", flush=True)
    print(f"SKIPPED (blank): {skipped_blank}", flush=True)
    print("First 5 inserted rows:")
    print(json.dumps(sample_rows, indent=2, default=str))

    cur.close()
    db.close()

    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    target = PROCESSED_DIR / SRC_PATH.name
    if target.exists():
        target.unlink()
    shutil.move(str(SRC_PATH), str(target))
    print(f"\nMoved xlsx -> {target.relative_to(_REPO_ROOT)}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
